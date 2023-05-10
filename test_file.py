import csv
import os

# TODO оформить в тест, добавить ассерты и использовать универсальный путь
def test_csv_file():
    PROJECT_ROOT_PATH = os.path.dirname(os.path.abspath(__file__))
    resources_package = os.path.join(PROJECT_ROOT_PATH, 'resources', 'eggs.csv')
    with open(resources_package, 'w') as csvfile:
       csvwriter = csv.writer(csvfile, delimiter=',')
       csvwriter.writerow(['Anna', 'Pavel', 'Peter'])
       csvwriter.writerow(['Alex', 'Serj', 'Yana'])

   with open(resources_package) as csvfile:
      csvreader = csv.reader(csvfile)
      for row in csvreader:
        print(row)

        import time

        from selenium import webdriver
        from selene import browser
        from selenium.webdriver.chrome.service import Service
        from webdriver_manager.chrome import ChromeDriverManager

        # TODO оформить в тест, добавить ассерты и использовать универсальный путь к tmp

        options = webdriver.ChromeOptions()
        prefs = {
            "download.default_directory": '/Users/kot/GitHubProjects/qa-guru/qa_guru_python_5_7_files',
            "download.prompt_for_download": False
        }
        options.add_experimental_option("prefs", prefs)

        browser.config.driver_options = options

        browser.open("https://github.com/pytest-dev/pytest")
        browser.element(".d-none .Button-label").click()
        browser.element('[data-open-app="link"]').click()

        import os.path

        import requests


        def test_downloaded_file_size():
            # TODO сохранять и читать из tmp, использовать универсальный путь
            url = 'https://selenium.dev/images/selenium_logo_square_green.png'

            r = requests.get(url)
            with open('selenium_logo.png', 'wb') as file:
                file.write(r.content)

            size = os.path.getsize('selenium_logo.png')

            assert size == 30803


from pypdf import PdfReader
# TODO оформить в тест, добавить ассерты и использовать универсальный путь
reader = PdfReader("resources/docs-pytest-org-en-latest.pdf")
number_of_pages = len(reader.pages)
page = reader.pages[0]
text = page.extract_text()
print(page)
print(number_of_pages)
print(text)

from openpyxl import load_workbook
# TODO оформить в тест, добавить ассерты и использовать универсальный путь

workbook = load_workbook('resources/file_example_XLSX_50.xlsx')
sheet = workbook.active
print(sheet.cell(row=3, column=2).value)

import xlrd
# TODO оформить в тест, добавить ассерты и использовать универсальный путь

book = xlrd.open_workbook('resources/file_example_XLS_10.xls')
print(f'Количество листов {book.nsheets}')
print(f'Имена листов {book.sheet_names()}')
sheet = book.sheet_by_index(0)
print(f'Количество столбцов {sheet.ncols}')
print(f'Количество строк {sheet.nrows}')
print(f'Пересечение строки 9 и столбца 1 = {sheet.cell_value(rowx=0, colx=1)}')
# печать всех строк по очереди
for rx in range(sheet.nrows):
    print(sheet.row(rx))
