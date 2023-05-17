import os
from openpyxl import load_workbook


# TODO оформить в тест, добавить ассерты и использовать универсальный путь
def test_xlsx():
    PROJECT_ROOT_PATH = os.path.dirname(os.path.abspath(__file__))
    resources_package = os.path.join(PROJECT_ROOT_PATH, '../resources', 'file_example_XLSX_50.xlsx')
    workbook = load_workbook(resources_package)
    sheet = workbook.active
    print(sheet.cell(row=3, column=2).value)
    assert sheet.cell(row=3, column=2).value == 'Mara'
